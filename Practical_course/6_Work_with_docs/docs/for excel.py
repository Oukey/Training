# -*- coding: utf-8 -*-
"""Работа с документами excel"""

# pip install openpyxl
from openpyxl import Workbook, load_workbook

# Создание пустого файла

wb = Workbook()
dest_filename = 'empty_book.xlsx'

# создание в документе нового листа с название Test
ws2 = wb.create_sheet(title="Test")

# запись значения в ячейку F5
ws2['F5'] = 3.14

# Запись в ячейку по указанной строке и столбцу
ws2.cell(column=5, row=4, value="12345")

# сщхранение нового документа в файле
wb.save(filename=dest_filename)

# загрузка существующего файла
wb = load_workbook(filename=dest_filename)

# Выбор нужного листа
ws = wb["Test"]

# Получение значение по названию ячейки
print(ws['F5'].value)
